import openpyxl
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox

# Absolute paths to your Excel files
PRODUCT_FILE_PATH = 'C:\\Users\\Hey!\\OneDrive\\Desktop\\Excel code\\product.xlsx'
CUSTOMER_FILE_PATH = 'C:\\Users\\Hey!\\OneDrive\\Desktop\\Excel code\\customer.xlsx'

def load_workbooks():
    global product_wb, product_ws, customer_wb, customer_ws
    try:
        product_wb = openpyxl.load_workbook(PRODUCT_FILE_PATH)
        product_ws = product_wb.active
        customer_wb = openpyxl.load_workbook(CUSTOMER_FILE_PATH)
        customer_ws = customer_wb.active
    except FileNotFoundError as e:
        print(f"Error loading file: {e}")
        messagebox.showerror("File Error", f"Error loading file: {e}")
        exit()
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
        messagebox.showerror("Error", f"An unexpected error occurred: {e}")
        exit()

def get_product_details(product_id, product_name):
    for row in product_ws.iter_rows(min_row=2, values_only=True):
        if row[0] == product_id and row[1] == product_name:
            return row
    return None

def save_customer_order(order_id, customer_name, product_id, product_name, quantity):
    product_details = get_product_details(product_id, product_name)
    if not product_details:
        messagebox.showerror("Product Not Found", "The specified product was not found.")
        return

    price = product_details[2]
    total_price = price * quantity

    # Append the order to the customer orders file
    customer_ws.append([order_id, customer_name, product_id, quantity, total_price])
    customer_wb.save(CUSTOMER_FILE_PATH)

    print(f"Order saved: {customer_name} ordered {quantity} of {product_name} for {total_price:.2f}")

    # Confirm order by showing the specific product row
    show_product_list(product_id, product_name)

def show_product_list(product_id, product_name):
    # Create a Tkinter window to display the product details
    window = tk.Tk()
    window.title("Product Details")

    tree = ttk.Treeview(window, columns=("ID", "Name", "Price"), show='headings')
    tree.heading("ID", text="Product ID")
    tree.heading("Name", text="Product Name")
    tree.heading("Price", text="Price")

    product_details = get_product_details(product_id, product_name)
    if product_details:
        tree.insert("", "end", values=product_details)
    else:
        messagebox.showinfo("No Product Found", "The specified product was not found.")

    tree.pack()

    # Add a button to close the window
    close_button = tk.Button(window, text="Close", command=window.destroy)
    close_button.pack()

    # Automatically close the window after a delay
    window.after(5000, window.destroy)  # Close after 5 seconds

    window.mainloop()

def submit_order(window):
    try:
        order_id = int(order_id_entry.get())
        customer_name = customer_name_entry.get().strip()
        product_id = int(product_id_entry.get())
        product_name = product_name_entry.get().strip()
        quantity = int(quantity_entry.get())

        if not customer_name or quantity <= 0 or not product_name:
            messagebox.showwarning("Invalid Input", "Please fill out all fields correctly.")
            return

        save_customer_order(order_id, customer_name, product_id, product_name, quantity)
        messagebox.showinfo("Success", "Order has been saved successfully.")

        # Close the Tkinter window
        window.destroy()
    except ValueError:
        messagebox.showerror("Invalid Input", "Please enter valid numbers for Order ID, Product ID, and Quantity.")

def order_input_form():
    window = tk.Tk()
    window.title("Enter Customer Order")

    # Create labels and entry widgets
    tk.Label(window, text="Order ID").grid(row=0, column=0)
    tk.Label(window, text="Customer Name").grid(row=1, column=0)
    tk.Label(window, text="Product ID").grid(row=2, column=0)
    tk.Label(window, text="Product Name").grid(row=3, column=0)
    tk.Label(window, text="Quantity").grid(row=4, column=0)

    global order_id_entry, customer_name_entry, product_id_entry, product_name_entry, quantity_entry
    order_id_entry = tk.Entry(window)
    customer_name_entry = tk.Entry(window)
    product_id_entry = tk.Entry(window)
    product_name_entry = tk.Entry(window)
    quantity_entry = tk.Entry(window)

    order_id_entry.grid(row=0, column=1)
    customer_name_entry.grid(row=1, column=1)
    product_id_entry.grid(row=2, column=1)
    product_name_entry.grid(row=3, column=1)
    quantity_entry.grid(row=4, column=1)

    # Create submit button
    submit_button = tk.Button(window, text="Submit Order", command=lambda: submit_order(window))
    submit_button.grid(row=5, columnspan=2)

    window.mainloop()

# Main execution
if __name__ == "__main__":
    load_workbooks()
    order_input_form()

